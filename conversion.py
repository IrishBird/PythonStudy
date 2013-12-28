import csv
from openpyxl import Workbook
from openpyxl.cell import get_column_letter


def cvsConversionExcel(filePath):

    #f = open('d:\\AAPL.csv', "rU")
    filePath="d:\\AAPL.csv"

    f=open(filePath,"rU")

    csv.register_dialect('singlequote', delimiter=',')
    #csv.register_dialect('colons', delimiter=':')

    reader = csv.reader(f, dialect='singlequote')

    wb = Workbook()
    dest_filename = "d:\\APPL2.xls"

    ws = wb.worksheets[0]

    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):
            column_letter = get_column_letter((column_index + 1))
            ws.cell('%s%s'%(column_letter, (row_index + 1))).value = cell

    wb.save(filename = dest_filename)

if __name__=="__main__":
    cvsConversionExcel("test")
      
 
